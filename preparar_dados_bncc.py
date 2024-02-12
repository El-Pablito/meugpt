from openpyxl import load_workbook
import json
import os

diretorio_planilhas = '/home/pablito/meu_gpt/planilhas'
planilha = 'bncc_ensino_fundamental.xlsx'
caminho_completo = os.path.join(diretorio_planilhas, planilha)

dados_treinamento = []

def extrair_dados(planilha):
    workbook = load_workbook(filename=planilha, data_only=True)
    for nome_aba in workbook.sheetnames:
        sheet = workbook[nome_aba]
        for row in sheet.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue  # Pula linhas inteiramente vazias
            # Substitua os índices [0] e [1] pelos índices das colunas de interesse
            prompt = f"Detalhes da linha na aba {nome_aba}."
            completion = f"{row[0]}, {row[1]}."
            dados_treinamento.append({'prompt': prompt, 'completion': completion})

extrair_dados(caminho_completo)

arquivo_saida = 'dados_treinamento_bncc.json'
with open(arquivo_saida, 'w', encoding='utf-8') as f:
    json.dump(dados_treinamento, f, ensure_ascii=False, indent=4)

print(f"Dados salvos em {arquivo_saida}")
